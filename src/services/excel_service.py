"""Creación del catálogo y modificación no destructiva del Excel opcional."""

from __future__ import annotations

from copy import copy
from io import BytesIO

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.domain.models import CatalogRow, ExcelSearchField
from src.services.matching_service import ConceptMatcher

CATALOG_COLUMNS = [
    "Clave del concepto",
    "Descripción completa",
    "Unidad",
    "Cantidad",
    "Precio unitario",
    "Estado de extracción",
    "Nivel de confianza",
]


class ExcelService:
    ORANGE = "FF5E12"
    ORANGE_LIGHT = "FF7D42"
    CHARCOAL = "362D32"
    BLUE = "00304F"

    def catalog_dataframe(self, catalog: list[CatalogRow]) -> pd.DataFrame:
        return pd.DataFrame(
            [
                {
                    "Clave del concepto": row.key,
                    "Descripción completa": row.description,
                    "Unidad": row.unit,
                    "Cantidad": row.quantity,
                    "Precio unitario": row.unit_price,
                    "Estado de extracción": row.status.value,
                    "Nivel de confianza": row.confidence / 100,
                }
                for row in catalog
            ],
            columns=CATALOG_COLUMNS,
        )

    def catalog_to_excel(self, catalog: list[CatalogRow]) -> bytes:
        dataframe = self.catalog_dataframe(catalog)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Catálogo extraído"

        for column_index, column_name in enumerate(dataframe.columns, start=1):
            cell = sheet.cell(row=1, column=column_index, value=column_name)
            cell.fill = PatternFill("solid", fgColor=self.BLUE)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_index, values in enumerate(dataframe.itertuples(index=False), start=2):
            for column_index, value in enumerate(values, start=1):
                sheet.cell(
                    row=row_index,
                    column=column_index,
                    value=self._excel_safe_value(value),
                )

        if len(dataframe) > 0:
            table = Table(displayName="CatalogoExtraido", ref=f"A1:G{len(dataframe) + 1}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)

        widths = {"A": 22, "B": 90, "C": 14, "D": 16, "E": 18, "F": 22, "G": 20}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        sheet.row_dimensions[1].height = 30
        for cell in sheet["D"][1:]:
            cell.number_format = '#,##0.0000'
        for cell in sheet["E"][1:]:
            cell.number_format = '$#,##0.00'
        for cell in sheet["G"][1:]:
            cell.number_format = '0%'
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def cross_reference_excel(
        self,
        excel_bytes: bytes,
        catalog: list[CatalogRow],
        matcher: ConceptMatcher,
        search_field: ExcelSearchField = ExcelSearchField.DESCRIPTION,
    ) -> tuple[bytes, pd.DataFrame]:
        workbook = load_workbook(BytesIO(excel_bytes))
        sheet = workbook.active
        search_column = 1 if search_field is ExcelSearchField.KEY else 2
        header_row = self._detect_header_row(sheet, search_field)
        output_column = sheet.max_column + 1
        header = sheet.cell(row=header_row, column=output_column, value="Precio unitario (PDF)")
        source_header = sheet.cell(row=header_row, column=max(1, output_column - 1))
        self._copy_header_style(source_header, header)
        if header.fill.fill_type is None:
            header.fill = PatternFill("solid", fgColor=self.BLUE)
            header.font = Font(color="FFFFFF", bold=True)

        preview_rows: list[dict[str, object]] = []
        for row_number in range(header_row + 1, sheet.max_row + 1):
            query = str(sheet.cell(row=row_number, column=search_column).value or "").strip()
            if not query:
                continue
            result = (
                matcher.match_key(query, catalog)
                if search_field is ExcelSearchField.KEY
                else matcher.match(query, catalog)
            )
            price_cell = sheet.cell(row=row_number, column=output_column)
            if result.status == "Encontrado" and result.unit_price is not None:
                price_cell.value = result.unit_price
                price_cell.number_format = '$#,##0.00'
            else:
                price_cell.value = None
            preview_rows.append(
                {
                    "Campo de búsqueda": search_field.value,
                    "Valor buscado": query,
                    "Precio unitario (PDF)": result.unit_price
                    if result.status == "Encontrado"
                    else None,
                    "Estado": result.status,
                    "Coincidencia": result.score / 100,
                }
            )

        sheet.column_dimensions[header.column_letter].width = 22
        output = BytesIO()
        workbook.save(output)
        return output.getvalue(), pd.DataFrame(preview_rows)

    @staticmethod
    def _detect_header_row(sheet, search_field: ExcelSearchField) -> int:
        column = 1 if search_field is ExcelSearchField.KEY else 2
        expected = ("clave", "codigo") if search_field is ExcelSearchField.KEY else (
            "concepto",
            "descrip",
        )
        for row_number in range(1, min(sheet.max_row, 20) + 1):
            value = str(sheet.cell(row=row_number, column=column).value or "").strip().lower()
            if any(term in value for term in expected):
                return row_number
        return 1

    @staticmethod
    def _copy_header_style(source, target) -> None:
        if source.has_style:
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)

    @staticmethod
    def _excel_safe_value(value):
        """Sustituye controles no válidos en XML sin alterar números ni celdas vacías."""
        if isinstance(value, str):
            return ILLEGAL_CHARACTERS_RE.sub(" ", value)
        return value
