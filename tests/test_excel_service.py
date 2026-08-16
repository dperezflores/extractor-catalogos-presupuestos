from io import BytesIO

from openpyxl import Workbook, load_workbook

from src.domain.models import CatalogRow, ExcelSearchField, ExtractionStatus
from src.services.excel_service import ExcelService
from src.services.matching_service import ConceptMatcher


def sample_catalog() -> list[CatalogRow]:
    return [
        CatalogRow(
            key="A-01",
            description="Suministro y colocación de concreto hidráulico",
            unit="m3",
            quantity=10,
            unit_price=1450.75,
            status=ExtractionStatus.EXTRACTED,
            confidence=100,
        )
    ]


def input_excel() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Clave"
    sheet["B1"] = "Concepto"
    sheet["A2"] = "A-01"
    sheet["B2"] = "Suministro y colocación de concreto hidráulico"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_catalog_excel_has_expected_columns() -> None:
    content = ExcelService().catalog_to_excel(sample_catalog())
    sheet = load_workbook(BytesIO(content)).active
    assert sheet["A1"].value == "Clave del concepto"
    assert sheet["G1"].value == "Nivel de confianza"
    assert sheet["H1"].value == "Ubicación en el PDF"
    assert sheet["H2"].value is None


def test_catalog_excel_reports_location_only_for_manual_review() -> None:
    catalog = sample_catalog()
    catalog[0].status = ExtractionStatus.REVIEW
    catalog[0].source_page_start = 17
    catalog[0].source_page_end = 20

    content = ExcelService().catalog_to_excel(catalog)
    sheet = load_workbook(BytesIO(content)).active

    assert sheet["H2"].value == "Páginas 17–20"


def test_catalog_excel_replaces_illegal_control_characters() -> None:
    catalog = sample_catalog()
    catalog[0].description = "Suministro y colocación\x0bde base hidráulica"
    catalog[0].unit = "m\x003"

    content = ExcelService().catalog_to_excel(catalog)
    sheet = load_workbook(BytesIO(content)).active

    assert sheet["B2"].value == "Suministro y colocación de base hidráulica"
    assert sheet["C2"].value == "m 3"


def test_optional_excel_gets_price_column() -> None:
    content, preview = ExcelService().cross_reference_excel(
        input_excel(), sample_catalog(), ConceptMatcher()
    )
    sheet = load_workbook(BytesIO(content)).active
    assert sheet["C1"].value == "Precio unitario (PDF)"
    assert sheet["C2"].value == 1450.75
    assert preview.iloc[0]["Estado"] == "Encontrado"


def test_optional_excel_can_search_by_key_column() -> None:
    content, preview = ExcelService().cross_reference_excel(
        input_excel(),
        sample_catalog(),
        ConceptMatcher(),
        ExcelSearchField.KEY,
    )
    sheet = load_workbook(BytesIO(content)).active
    assert sheet["C2"].value == 1450.75
    assert preview.iloc[0]["Campo de búsqueda"] == "Clave (columna A)"
    assert preview.iloc[0]["Valor buscado"] == "A-01"
